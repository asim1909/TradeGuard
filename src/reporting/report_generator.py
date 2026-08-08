"""
Report Generator Module for Trade Reconciliation & Control Automation Engine.

Reads reconciliation audit records from SQLite database and generates executive
Excel workbooks (.xlsx), raw CSV exports, and structured JSON audit files.
"""

from datetime import datetime
from pathlib import Path
import time
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import CSV_REPORTS_DIR, EXCEL_REPORTS_DIR, JSON_REPORTS_DIR, REPORTS_DIR
from src.database.database_manager import DatabaseManager
from src.utils.exceptions import DatabaseConnectionError, ReportGenerationError
from src.utils.logger import get_logger

logger = get_logger("ReportGenerator")


class ReportGenerator:
    """
    Manages generation of OpenPyXL Excel workbooks, CSV feeds, and JSON audit files from SQLite data.

    Attributes:
        db_manager: DatabaseManager instance for querying SQLite tables.
        reports_dir: Base reports folder path.
    """

    SEVERITY_STYLES = {
        "CRITICAL": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                     "font": Font(name="Calibri", size=11, bold=True, color="9C0006")},
        "HIGH":     {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                     "font": Font(name="Calibri", size=11, bold=True, color="9C6500")},
        "MEDIUM":   {"fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                     "font": Font(name="Calibri", size=11, bold=True, color="7F6000")},
        "LOW":      {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                     "font": Font(name="Calibri", size=11, bold=True, color="006100")},
    }

    HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    def __init__(
        self,
        db_manager: Optional[DatabaseManager] = None,
        reports_dir: Path = REPORTS_DIR,
    ) -> None:
        """Initializes ReportGenerator with database connection and output paths."""
        self.db_manager = db_manager or DatabaseManager()
        self.reports_dir = Path(reports_dir) if isinstance(reports_dir, str) else reports_dir
        self.excel_dir = self.reports_dir / "excel"
        self.csv_dir = self.reports_dir / "csv"
        self.json_dir = self.reports_dir / "json"
        self._ensure_output_directories()
        logger.info(f"Initialized ReportGenerator targeting base directory: {self.reports_dir}")

    def _ensure_output_directories(self) -> None:
        """Creates target destination folders for Excel, CSV, and JSON outputs."""
        self.excel_dir.mkdir(parents=True, exist_ok=True)
        self.csv_dir.mkdir(parents=True, exist_ok=True)
        self.json_dir.mkdir(parents=True, exist_ok=True)

    def read_reconciliation_data(self) -> Tuple[Dict[str, Any], pd.DataFrame]:
        """Reads latest reconciliation run summary and break records from SQLite database."""
        logger.info("Reading reconciliation run data from SQLite database...")
        self.db_manager.connect()

        if not self.db_manager.table_exists("reconciliation_summary"):
            raise ReportGenerationError("Table 'reconciliation_summary' missing. Run 'reconcile' first.")

        summary_rows = self.db_manager.execute_select(
            "SELECT * FROM reconciliation_summary ORDER BY Created_At DESC LIMIT 1"
        )
        if not summary_rows:
            raise ReportGenerationError("No reconciliation run history found in database. Run 'reconcile' first.")

        summary_dict = summary_rows[0]
        breaks_df = self.db_manager.fetch_dataframe("SELECT * FROM reconciliation_breaks")

        logger.info(f"Retrieved run summary '{summary_dict.get('Run_ID')}' and {len(breaks_df)} break records.")
        return summary_dict, breaks_df

    def generate_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        summary: Dict[str, Any],
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Generates formatted Executive Summary sheet with KPI cards and metrics table."""
        ws = wb.create_sheet(title="Executive Summary")
        ws.views.sheetView[0].showGridLines = True

        # Title Header
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "TRADE RECONCILIATION EXECUTIVE CONTROL SUMMARY"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1B365D")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # KPI Metrics Table
        data_rows = [
            ("Run ID", summary.get("Run_ID", "N/A")),
            ("Execution Time", f"{float(summary.get('Execution_Time', 0.0)):.4f} seconds"),
            ("Generated On", str(summary.get("Created_At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))),
            ("Front Office Trades", int(summary.get("Front_Count", 0))),
            ("Back Office Trades", int(summary.get("Back_Count", 0))),
            ("Matched Trades", int(summary.get("Matched_Count", 0))),
            ("Match Percentage", f"{float(summary.get('Match_Percentage', 0.0)):.2f}%"),
            ("Critical Issues", int(summary.get("Critical_Breaks", 0))),
            ("High Issues", int(summary.get("High_Breaks", 0))),
            ("Medium Issues", int(summary.get("Medium_Breaks", 0))),
            ("Low Issues", int(summary.get("Low_Breaks", 0))),
        ]

        ws.append([])
        ws.append(["Metric Attribute", "Value"])
        ws.row_dimensions[3].height = 24

        for col_idx in range(1, 3):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for r_idx, (attr, val) in enumerate(data_rows, start=4):
            ws.append([attr, val])
            ws.cell(row=r_idx, column=1).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(row=r_idx, column=1).border = self.THIN_BORDER
            val_cell = ws.cell(row=r_idx, column=2)
            val_cell.border = self.THIN_BORDER
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            if r_idx % 2 == 1:
                ws.cell(row=r_idx, column=1).fill = self.ZEBRA_FILL
                val_cell.fill = self.ZEBRA_FILL

        ws.freeze_panes = "A4"
        self._auto_fit_column_widths(ws)
        return ws

    def generate_exception_sheet(
        self,
        wb: openpyxl.Workbook,
        sheet_title: str,
        df_subset: pd.DataFrame,
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Generates individual styled worksheet for a specific break category."""
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        headers = ["Trade_ID", "Break_Type", "Expected_Value", "Actual_Value", "Severity", "Detected_At"]
        ws.append(headers)
        ws.row_dimensions[1].height = 24

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if not df_subset.empty:
            for r_idx, row in enumerate(df_subset.to_dict("records"), start=2):
                row_vals = [row.get(h, "") for h in headers]
                ws.append(row_vals)
                severity = str(row.get("Severity", "MEDIUM")).upper()

                for c_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = self.THIN_BORDER
                    cell.alignment = Alignment(horizontal="center" if c_idx in (1, 5, 6) else "left", vertical="center")
                    if r_idx % 2 == 1:
                        cell.fill = self.ZEBRA_FILL

                # Apply conditional severity highlight
                if severity in self.SEVERITY_STYLES:
                    sev_cell = ws.cell(row=r_idx, column=5)
                    sev_cell.fill = self.SEVERITY_STYLES[severity]["fill"]
                    sev_cell.font = self.SEVERITY_STYLES[severity]["font"]

        ws.freeze_panes = "A2"
        self._auto_fit_column_widths(ws)
        return ws

    def format_workbook(self, wb: openpyxl.Workbook) -> None:
        """Applies global formatting, column widths, and gridline settings across all sheets."""
        for ws in wb.worksheets:
            ws.views.sheetView[0].showGridLines = True
            self._auto_fit_column_widths(ws)

    def generate_excel_report(
        self,
        summary_dict: Dict[str, Any],
        breaks_df: pd.DataFrame,
    ) -> Path:
        """Generates OpenPyXL 10-sheet Excel workbook report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_path = self.excel_dir / f"Trade_Reconciliation_Report_{timestamp}.xlsx"
        logger.info(f"Generating Excel report workbook at: {target_path}")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default active sheet

        # 1. Executive Summary Sheet
        self.generate_summary_sheet(wb, summary_dict)

        # 2-10. Exception Break Sheets
        sheet_mapping = [
            ("Missing Trades", "Missing Trade"),
            ("Unexpected Trades", "Unexpected Trade"),
            ("Price Mismatches", "Price Mismatch"),
            ("Quantity Mismatches", "Quantity Mismatch"),
            ("Currency Mismatches", "Currency Mismatch"),
            ("Status Mismatches", "Status Mismatch"),
            ("Trade Date Mismatches", "Trade Date Mismatch"),
            ("Settlement Date Mismatches", "Settlement Date Mismatch"),
            ("Duplicate Trades", "Duplicate Trade"),
        ]

        for sheet_title, break_type in sheet_mapping:
            if not breaks_df.empty and "Break_Type" in breaks_df.columns:
                df_sub = breaks_df[breaks_df["Break_Type"] == break_type].copy()
            else:
                df_sub = pd.DataFrame()
            self.generate_exception_sheet(wb, sheet_title, df_sub)

        self.format_workbook(wb)
        wb.save(target_path)
        logger.info(f"Successfully saved Excel report: {target_path}")
        return target_path

    def generate_csv_reports(
        self,
        summary_dict: Dict[str, Any],
        breaks_df: pd.DataFrame,
    ) -> List[Path]:
        """Generates individual CSV break reports and summary.csv in reports/csv/."""
        logger.info("Generating CSV break reports...")
        saved_paths: List[Path] = []

        # 1. summary.csv
        sum_path = self.csv_dir / "summary.csv"
        pd.DataFrame([summary_dict]).to_csv(sum_path, index=False, encoding="utf-8")
        saved_paths.append(sum_path)

        # 2. Category CSV files
        csv_mapping = [
            ("missing_trades.csv", "Missing Trade"),
            ("unexpected_trades.csv", "Unexpected Trade"),
            ("price_mismatches.csv", "Price Mismatch"),
            ("quantity_mismatches.csv", "Quantity Mismatch"),
            ("currency_mismatches.csv", "Currency Mismatch"),
            ("status_mismatches.csv", "Status Mismatch"),
            ("duplicate_trades.csv", "Duplicate Trade"),
        ]

        for filename, break_type in csv_mapping:
            path = self.csv_dir / filename
            if not breaks_df.empty and "Break_Type" in breaks_df.columns:
                sub_df = breaks_df[breaks_df["Break_Type"] == break_type]
            else:
                sub_df = pd.DataFrame(columns=["Trade_ID", "Break_Type", "Expected_Value", "Actual_Value", "Severity", "Detected_At"])
            sub_df.to_csv(path, index=False, encoding="utf-8")
            saved_paths.append(path)

        logger.info(f"Successfully generated {len(saved_paths)} CSV report files in {self.csv_dir}")
        return saved_paths

    def generate_json_reports(
        self,
        summary_dict: Dict[str, Any],
        breaks_df: pd.DataFrame,
    ) -> Tuple[Path, Path]:
        """Generates summary.json and breaks.json audit files in reports/json/."""
        logger.info("Generating JSON audit report files...")
        sum_json_path = self.json_dir / "summary.json"
        breaks_json_path = self.json_dir / "breaks.json"

        # Write summary.json
        import json
        with open(sum_json_path, "w", encoding="utf-8") as f:
            json.dump(summary_dict, f, indent=4, default=str)

        # Write breaks.json
        breaks_list = breaks_df.to_dict("records") if not breaks_df.empty else []
        with open(breaks_json_path, "w", encoding="utf-8") as f:
            json.dump(breaks_list, f, indent=4, default=str)

        logger.info(f"Successfully generated JSON audit files: {sum_json_path}, {breaks_json_path}")
        return sum_json_path, breaks_json_path

    def run(self) -> Dict[str, Any]:
        """Executes full reporting pipeline and exports Excel, CSV, and JSON files."""
        start_time = time.perf_counter()
        logger.info("Starting enterprise report generation workflow...")

        summary_dict, breaks_df = self.read_reconciliation_data()

        excel_path = self.generate_excel_report(summary_dict, breaks_df)
        csv_paths = self.generate_csv_reports(summary_dict, breaks_df)
        json_paths = self.generate_json_reports(summary_dict, breaks_df)

        elapsed = round(time.perf_counter() - start_time, 4)
        logger.info(f"Report generation workflow completed in {elapsed} seconds.")

        return {
            "excel": excel_path,
            "csv": csv_paths,
            "json": json_paths,
            "reports_dir": self.reports_dir,
            "elapsed_time": elapsed,
        }

    def _auto_fit_column_widths(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Autofits worksheet column widths based on maximum string length."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if cell.coordinate in ws.merged_cells:
                    continue
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
